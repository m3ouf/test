from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from database_backend.models import MSAN, Router, RouterNetwork, RouterPort, TEDataMsan
from collections import namedtuple
import logging
import sys
import re

logger = logging.getLogger(__name__)
CNT_THREASHOLD = 1
error_log = open("/tmp/tedata_msans_errors", "w")


def reformat_integer(pattern):
    return "{0}-{1}-{2}-{3}".format(pattern.group(1), int(pattern.group(2)), pattern.group(3), pattern.group(4))


class Command(BaseCommand):
    def handle(self, *args, **options):
        logger.info("*** Script started ***")
        logger.info("*** Parsing file ***")
        if not len(sys.argv) == 3:
            raise CommandError("You need to specify a valid xlsx file\n"
            "usage: python manage.py migrate_te_msan_plans sample.xlsx")
        wb = load_workbook(sys.argv[2])
        ws = wb.active
        integer_pattern = re.compile(r'^(\d{2})-(\d{2})-(\d{2})-(\d{2})')

        count = 1
        latest_router = None
        latest_subnet = None
        for i in range(3, len(ws.rows), 2):
            if count % CNT_THREASHOLD == 0:
                logger.info("reached MSAN %s", str(count))
            try:
                msan_code = integer_pattern.sub(reformat_integer, ws.cell(row=i, column=2).value)

                cells = [ws.cell(row=i, column=j).value for j in range(15, 22)]
                cells[1] = cells[1] or latest_subnet
                latest_subnet = cells[1]
                if not cells[1].endswith('/24'):
                    cells[1] += '/24'

                shelf1 = cells[4]
                shelf2 = ws.cell(row=i+1, column=19).value
                shelf1 = re.sub(r"\s*shelf\s*\d\s*\:\s*(.*)\s*", r"\1", shelf1, flags=re.IGNORECASE)
                shelf2 = re.sub(r"\s*shelf\s*\d\s*\:\s*(.*)\s*", r"\1", shelf2, flags=re.IGNORECASE)
                cells[4] = shelf1
                cells.insert(5, shelf2)
                cells[-2] = cells[-2] or latest_router

                latest_router = cells[-2]

                TeDataPlan = namedtuple('TeDataPlan', ['vlan', 'subnet', 'gw_subnet', 'gw_ip', 'shelf1', 'shelf2',
                                                       'router', 'port'])
                plan = TeDataPlan(*cells)

                msan = MSAN.objects.get(code=msan_code.strip())
                router, created = Router.objects.get_or_create(name=plan.router.strip())
                if created:  # router didn't exist before .. so it doesn't have ports .. let's create new ones !
                    router_port = RouterPort.objects.create(name=plan.port.strip(), used=False, router=router)
                    RouterNetwork.objects.create(network_ip=plan.subnet.strip(), router=router)
                else:
                    # we need to make sure that port is not already in use
                    if router.ports.filter(name=plan.port.strip()).count():
                        logger.error("Port %s of router %s is already assigned", plan.port, router.name)
                        count += 1
                        continue
                    else:
                        router_port = RouterPort.objects.create(name=plan.port.strip(), used=False, router=router)
                TEDataMsan.objects.create(msan=msan, router_port=router_port, manage_vlan=plan.vlan or 400,
                                          manage_gw_subnet=plan.gw_subnet.strip(), manage_gw_ip=plan.gw_ip.strip(),
                                          shelf1=plan.shelf1.strip(),
                                          shelf2=shelf2.strip())
            except MSAN.DoesNotExist:
                logger.error("MSAN with code %s doesn't exist", msan_code)
                error_log.write(msan_code + ": MSAN with code %s doesn't exist\n" % msan_code)

            except Exception as e:
                logger.error("Error while provisioning TeData Plan for MSAN %s, %s", msan_code, str(e))
                error_log.write(msan_code + ": Error while provisioning TeData Plan for MSAN %s, %s\n" %
                                (msan_code, str(e)))
            finally:
                count += 1

        error_log.close()
        logger.info("*** Script finished ***")