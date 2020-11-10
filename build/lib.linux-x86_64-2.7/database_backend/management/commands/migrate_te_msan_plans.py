from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from database_backend.models import MSAN
from django.core.exceptions import ValidationError
import logging
import sys
import re

logger = logging.getLogger(__name__)
CNT_THREASHOLD = 1
error_log = open("/tmp/te_msans_errors", "w")

def reformat_integer(pattern):
    return "{0}-{1}-{2}-{3}".format(pattern.group(1), int(pattern.group(2)), pattern.group(3), pattern.group(4))


class Command(BaseCommand):
    def handle(self, *args, **options):
        logger.info("*** Script started ***")
        logger.info("*** Parsing file ***")
        if not len(sys.argv) == 3:
            raise CommandError("You need to specify a valid xlsx file\n"
            "usage: python manage.py migrate_te_msan_plans sample.xlsx")
        wb = load_workbook(sys.argv[2])
        ws = wb.active
        fields = ['code', 'name', 'h248_subnet', 'gateway_interface', 'traffic_shelf1', 'traffic_shelf2',
                  'traffic_vlan', 'manage_subnet', 'manage_gw_subnet', 'manage_gw_ip', 'manage_shelf1',
                  'manage_shelf2', 'manage_vlan']
        integer_pattern = re.compile(r'^(\d{1,2})-(\d{1,2})-(\d{1,2})-(\d{1,2})')

        count = 1
        for i in range(3, len(ws.rows), 2):
            if count % CNT_THREASHOLD == 0:
                logger.info("reached MSAN %s", str(count))
            try:
                cells = [ws.cell(row=i, column=j).value for j in range(2, 14) if j != 6]

                if cells[0] is None or integer_pattern.search(cells[0]) is None:
                    raise ValidationError("Invalid MSAN code format")

                cells[0] = integer_pattern.sub(reformat_integer, cells[0])

                try:
                    int(cells[5])
                except ValueError:
                    raise ValidationError("Traffic VLAN must field must contain an integer value")

                cells[9] = re.sub(r"\s*shelf\s*\d\s*\:\s*(.*)\s*", r"\1", cells[9], flags=re.IGNORECASE)
                m_shelf2 = re.sub(r"\s*shelf\s*\d\s*\:\s*(.*)\s*", r"\1", ws.cell(row=i+1, column=12).value,
                                  flags=re.IGNORECASE)
                t_shelf2 = ws.cell(row=i+1, column=7).value
                cells.insert(10, m_shelf2)
                cells.insert(5, t_shelf2)
                cells[2] = re.sub("\s*", "", cells[2])
                cells[3] = re.sub("\s*", "", cells[3])
                cells[4] = re.sub("\s*", "", cells[4])
                cells[5] = re.sub("\s*", "", cells[5])
                cells[7] = re.sub("\s*", "", cells[7])
                cells[8] = re.sub("\s*", "", cells[8])
                msan = dict(zip(fields, cells))

                MSAN.objects.create(**msan)

            except Exception as e:
                logger.error("Error while provisioning Te Plan for MSAN %s, %s", str(cells[0]), str(e))
                error_log.write(str(cells[0]) + ": Error while provisioning Te Plan for MSAN %s, %s\n" %
                                (str(cells[0]), str(e)))

                
            finally:
                 count += 1
        error_log.close()
        logger.info("*** Script finished ***")