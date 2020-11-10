#!/usr/bin/env python
from django.core.management.base import BaseCommand, CommandError
from physical_adsl_backend.clients import get_client
from physical_adsl_backend.behaviors import HuaweiTelnetBehavior
from Exscript.protocols.Exception import InvalidCommandException

import logging
import sys
from openpyxl import load_workbook
logger = logging.getLogger(__name__)


class Command(BaseCommand):
    def handle(self, *args, **options):
        logger.info("*** Script started ***")
        if not len(sys.argv) == 3:
            raise CommandError("You need to specify a valid xlsx file\n"
            "usage: python manage.py configure_batch_pilot sample.xlsx")

        wb = load_workbook(sys.argv[2])

        with open("/tmp/pilot_batch_errors.txt", 'w', 1) as errors:
            logger.info("Starting to process pilot")
            for sheet_name in wb.get_sheet_names():
                logger.info("Starting to process %s ips", sheet_name)
                wd = wb[sheet_name]
                for i in range(2, len(wd.rows) + 1):
                    try:
                        ip = wd.cell(row=i, column=1).value.strip()
                        card = wd.cell(row=i, column=2).value
                        port = wd.cell(row=i, column=3).value
                        logger.info("handling IP %s, %s/%s", ip, card, port)

                        if sheet_name == "Hu5600":
                            with HuaweiTelnetBehavior(ip, '5600') as h5600telnet_bhavior:
                                try:
                                    h5600telnet_bhavior.execute("config")
                                    h5600telnet_bhavior.execute("traffic table index 7 ip car 24576 priority 7 priority-policy pvc-Setting")
                                except InvalidCommandException as e:
                                    if "Failed to create traffic descriptor record,the index has existed" in str(e):
                                        pass
                                    else:
                                        raise
                        # elif sheet_name == "Hu5600T":
                        #     h5600Ttelnet_behavior.set_port_info(ip, "traffic table ip index 7 cir 24576 priority 7 inner-priority 7 priority-policy tag-In-package")
                        client = get_client(sheet_name)

                        result = client.provision_subscriber(ip, card, port)
                        if not result:
                            errors.write(ip + ": Error while provisioning\n")

                    except Exception as e:
                        errors.write(ip + ": " + str(e) + "\n")

            logger.info("Finished to process pilot")
        logger.info("*** Script finished ***")
