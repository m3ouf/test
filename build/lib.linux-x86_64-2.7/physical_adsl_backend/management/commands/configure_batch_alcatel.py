#!/usr/bin/env python
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from physical_adsl_backend.behaviors import AlcatelTelnetBehavior
import logging
import sys
logger = logging.getLogger(__name__)
CNT_THREASHOLD = 10000


class Command(BaseCommand):
    def handle(self, *args, **options):
        logger.info("*** Script started ***")
        if not len(sys.argv) == 3:
            raise CommandError("You need to specify a valid xlsx file\n"
            "usage: python manage.py configure_batch_alcatel sample.xlsx")
        wb = load_workbook(sys.argv[2])
        telnet_behavior = AlcatelTelnetBehavior()

        with open("/tmp/alcatel_batch_errors.txt", 'w') as errors:

            xd = wb.get_sheet_by_name("XD")
            fd = wb.get_sheet_by_name("FD")

            # handle XD ips
            logger.info("Starting to process XD ips")
            for i in range(2, len(xd.rows) + 1):
                try:
                    ip = xd.cell(row=i, column=1).value.strip()
                    logger.info("handling IP %s", ip)

                    telnet_behavior.set_port_info(ip, "configure qos shub tc-map-dot1p 3 traffic-class 0")
                    telnet_behavior.set_port_info(ip, "configure qos shub tc-map-dot1p 0 traffic-class 3")
                    telnet_behavior.set_port_info(ip, "configure qos tc-map-dot1p 0 queue 3")
                    telnet_behavior.set_port_info(ip, "configure qos tc-map-dot1p 3 queue 0")
                except Exception as e:
                    errors.write("xd - " + ip + ": " + str(e) + "\n")

            logger.info("Finished to process XD ips")

            # handle FD ips
            logger.info("Starting to process FD ips")
            for i in range(2, len(fd.rows) + 1):
                try:
                    ip = fd.cell(row=i, column=1).value.strip()
                    logger.info("handling IP %s", ip)

                    telnet_behavior.set_port_info(ip, "configure qos shub tc-map-dot1p 0 traffic-class 3")
                    telnet_behavior.set_port_info(ip, "configure qos shub tc-map-dot1p 3 traffic-class 0")
                    telnet_behavior.set_port_info(ip, "configure qos tc-map-dot1p 0 tc 3")
                    telnet_behavior.set_port_info(ip, "configure qos tc-map-dot1p 3 tc 0")

                except Exception as e:
                    errors.write("fd - " + ip + ": " + str(e) + "\n")

            logger.info("Finished to process FD ips")
        logger.info("*** Script finished ***")
