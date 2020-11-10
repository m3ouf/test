from openpyxl.workbook import Workbook
from database_backend.models import MSAN
CNT_THREASHOLD = 1
SHELF1 = "Shelf1"
SHELF2 = "Shelf2"
BB_RANGE = "401-480"

def get_msan_plans_reports():
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        header1 = ["Traffic H248", None, None, None, None, None, None, None, "Management", None, None, None, None,
                   "TE Data", None, None, None, None, None, None, None]
        header2 = [None, "Code", "MSAN", "H.248 subnet", "Gateway Interface", "Shelf", "New Sig  & RTP", "Vlan",
                   "Management subnet", "Manag. Gateway Subnet", "Manag. Gateway IP", "management IP", "manag. VLAN",
                   "BB  Vlan", "manag.  TED Vlan", "Management subnet", "Manag. Gateway Subnet", "Manag. Gateway IP",
                   "management IP", "Router", "Port"]

        rows = [header1, header2]
        for i in range(len(rows)):
            for j in range(len(rows[i])):
                ws.cell(row=i + 1, column=j + 1).value = rows[i][j]

        msans = MSAN.objects.all()

        i += 2
        ids_count = 1
        for msan in msans:
            ws['A{}'.format(i)] = ids_count
            ws.merge_cells(start_row=i, start_column=1, end_row=i+1, end_column=1)
            ws['B{}'.format(i)] = msan.code
            ws.merge_cells(start_row=i, start_column=2, end_row=i+1, end_column=2)
            ws['C{}'.format(i)] = msan.name
            ws.merge_cells(start_row=i, start_column=3, end_row=i+1, end_column=3)
            ws['D{}'.format(i)] = str(msan.h248_subnet)
            ws.merge_cells(start_row=i, start_column=4, end_row=i+1, end_column=4)
            ws['E{}'.format(i)] = msan.gateway_interface
            ws.merge_cells(start_row=i, start_column=5, end_row=i+1, end_column=5)
            ws['F{}'.format(i)] = SHELF1
            ws['F{}'.format(i + 1)] = SHELF2
            ws['G{}'.format(i)] = msan.traffic_shelf1
            ws['G{}'.format(i + 1)] = msan.traffic_shelf2
            ws['H{}'.format(i)] = msan.traffic_vlan
            ws.merge_cells(start_row=i, start_column=8, end_row=i+1, end_column=8)
            ws['I{}'.format(i)] = str(msan.manage_subnet)
            ws.merge_cells(start_row=i, start_column=9, end_row=i+1, end_column=9)
            ws['J{}'.format(i)] = str(msan.manage_gw_subnet)
            ws.merge_cells(start_row=i, start_column=10, end_row=i+1, end_column=10)
            ws['K{}'.format(i)] = msan.manage_gw_ip
            ws.merge_cells(start_row=i, start_column=11, end_row=i+1, end_column=11)
            ws['L{}'.format(i)] = "{}: {}".format(SHELF1, msan.manage_shelf1)
            ws['L{}'.format(i + 1)] = "{}: {}".format(SHELF2, msan.manage_shelf2)
            ws['M{}'.format(i)] = msan.manage_vlan
            ws.merge_cells(start_row=i, start_column=13, end_row=i+1, end_column=13)
            ws['N{}'.format(i)] = BB_RANGE
            ws.merge_cells(start_row=i, start_column=14, end_row=i+1, end_column=14)

            if hasattr(msan, 'tedatamsan'):
                ws["O{}".format(i)] = msan.tedatamsan.manage_vlan
                ws.merge_cells(start_row=i, start_column=15, end_row=i+1, end_column=15)
                ws['P{}'.format(i)] = str(msan.tedatamsan.router_port.router.networks.all()[0])
                ws.merge_cells(start_row=i, start_column=16, end_row=i+1, end_column=16)
                ws['Q{}'.format(i)] = str(msan.tedatamsan.manage_gw_subnet)
                ws.merge_cells(start_row=i, start_column=17, end_row=i+1, end_column=17)
                ws['R{}'.format(i)] = msan.tedatamsan.manage_gw_ip
                ws.merge_cells(start_row=i, start_column=18, end_row=i+1, end_column=18)
                ws['S{}'.format(i)] = "{}: {}".format(SHELF1, msan.tedatamsan.shelf1)
                ws['S{}'.format(i + 1)] = "{}: {}".format(SHELF1, msan.tedatamsan.shelf2)
                ws['T{}'.format(i)] = msan.tedatamsan.router_port.router.name
                ws.merge_cells(start_row=i, start_column=20, end_row=i+1, end_column=20)
                ws['U{}'.format(i)] = msan.tedatamsan.router_port.name
                ws.merge_cells(start_row=i, start_column=21, end_row=i+1, end_column=21)

            i += 2
            ids_count += 1

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=13)
        ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=21)
        return wb